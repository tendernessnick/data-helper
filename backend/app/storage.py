"""数据集本地存储。

每个数据集位于 data/datasets/{id}/，包含：
- meta.json            元信息（名称、行列数、列类型、操作历史）
- original.*           用户上传的原始文件（用于回滚）
- current.parquet      当前工作副本（Parquet，仅由本程序写入和读取）
- prev.parquet         撤销快照（只保留最近一步）

历史版本使用 pickle（current.pkl），加载时自动迁移到 Parquet。
"""
import io
import json
import logging
import os
import shutil
import threading
import uuid
from datetime import datetime
from pathlib import Path

import pandas as pd

from .paths import DATA_DIR

logger = logging.getLogger(__name__)

DATASETS_DIR = DATA_DIR / "datasets"
DATASETS_DIR.mkdir(parents=True, exist_ok=True)

MAX_HISTORY = 200
# 原始文件超过该阈值时走流式建集（分块读 CSV 直接写 Parquet，避免整表进内存）
STREAM_THRESHOLD_BYTES = 16 * 1024 * 1024
STREAM_CHUNK_ROWS = 200_000
_lock = threading.RLock()


class DatasetNotFound(KeyError):
    pass


def _ds_dir(ds_id: str) -> Path:
    d = DATASETS_DIR / str(ds_id)
    if not d.is_dir():
        raise DatasetNotFound(ds_id)
    return d


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _read_meta(d: Path) -> dict:
    return json.loads((d / "meta.json").read_text(encoding="utf-8"))


def _write_meta(d: Path, meta: dict) -> None:
    tmp = d / "meta.json.tmp"
    tmp.write_text(json.dumps(meta, ensure_ascii=False, indent=1), encoding="utf-8")
    os.replace(tmp, d / "meta.json")  # 原子替换：写一半崩溃不让数据集从列表消失


def _columns_info(df: pd.DataFrame) -> list:
    return [{"name": str(c), "dtype": str(df[c].dtype)} for c in df.columns]


def _write_parquet(df: pd.DataFrame, path: Path) -> None:
    """原子写入 Parquet；病态 object 列（混合类型）退化为字符串后重试。"""
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        df.to_parquet(tmp)
    except Exception as e:
        logger.warning("Parquet 直写失败（%s），对 object 列做字符串化重试：%s", type(e).__name__, e)
        out = df.copy()
        for c in out.columns:
            if out[c].dtype == object:
                out[c] = out[c].astype(str).mask(out[c].isna(), None)
        out.to_parquet(tmp)
    os.replace(tmp, path)  # 写一半崩溃不留损坏的 current


def _migrate_legacy(d: Path) -> None:
    """旧版 pickle 存储自动迁移：current.pkl → current.parquet（调用方需持 _lock）。"""
    pkl = d / "current.pkl"
    if not pkl.exists():
        return
    df = pd.read_pickle(pkl)
    _write_parquet(df, d / "current.parquet")
    pkl.unlink(missing_ok=True)  # 并发首载时另一线程可能已迁移
    (d / "prev.pkl").unlink(missing_ok=True)  # 旧快照随迁移一并废弃
    logger.info("数据集 %s 已从 pickle 迁移到 Parquet", d.name)


def current_path(ds_id: str) -> Path:
    """当前工作副本的 Parquet 路径（必要时先迁移旧 pickle）。"""
    with _lock:  # 与 save_df 的 os.replace 互斥：Windows 上替换正被读取的文件会 PermissionError
        d = _ds_dir(ds_id)
        p = d / "current.parquet"
        if not p.exists():
            _migrate_legacy(d)
        if not p.exists():
            raise DatasetNotFound(ds_id)
        return p


def _decode_bytes(raw: bytes) -> str:
    for enc in ("utf-8-sig", "gbk"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1")


def read_csv_any(src) -> pd.DataFrame:
    """自动尝试常见编码和分隔符读取 CSV（src 为 bytes 或路径）。"""
    raw = src if isinstance(src, (bytes, bytearray)) else Path(src).read_bytes()
    text = _decode_bytes(bytes(raw))
    try:
        df = pd.read_csv(io.StringIO(text), sep=None, engine="python")
        # 嗅探在单列等场景会误判分隔符（整表变成 Unnamed 列），回退标准逗号解析
        if df.shape[1] and all(str(c).startswith("Unnamed:") for c in df.columns):
            return pd.read_csv(io.StringIO(text))
        return df
    except pd.errors.ParserError:
        # 分隔符嗅探失败时按标准逗号解析
        return pd.read_csv(io.StringIO(text))


def _parse_json(raw: bytes) -> pd.DataFrame:
    text = raw.decode("utf-8-sig")
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return pd.read_json(io.StringIO(text))
    if isinstance(data, list):
        if data and all(isinstance(x, dict) for x in data):
            return pd.json_normalize(data)
        return pd.DataFrame(data)
    if isinstance(data, dict):
        for key in ("data", "rows", "records", "items", "list"):
            if isinstance(data.get(key), list):
                inner = data[key]
                if inner and all(isinstance(x, dict) for x in inner):
                    return pd.json_normalize(inner)
                return pd.DataFrame(inner)
        return pd.DataFrame(data)
    raise ValueError("JSON 顶层必须是数组或对象")


def parse_upload(filename: str, raw: bytes, sheet_name=None) -> pd.DataFrame:
    ext = Path(filename or "").suffix.lower()
    if ext in (".csv", ".txt"):
        df = read_csv_any(raw)
    elif ext == ".xlsx":
        df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet_name if sheet_name is not None else 0)
    elif ext == ".json":
        df = _parse_json(raw)
    else:
        raise ValueError(f"不支持的文件类型 {ext or '(无后缀)'}，请上传 CSV / XLSX / JSON")
    if df.empty:
        raise ValueError("文件解析后没有任何数据行")
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [" ".join(str(x) for x in tup if str(x) != "nan") for tup in df.columns]
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _new_ds_id() -> str:
    """生成数据集目录 id：时间戳 + uuid6 位。占用即重生成（防碰撞/TOCTOU）。"""
    for _ in range(8):
        ds_id = datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:6]
        if not (DATASETS_DIR / ds_id).exists():
            return ds_id
    raise RuntimeError("无法生成可用的数据集 id（目录异常，请检查 data/datasets）")


def create_dataset(name, df: pd.DataFrame, original_filename: str, original_bytes: bytes,
                   action: str = "上传", detail: str = "") -> str:
    ds_id = _new_ds_id()
    with _lock:
        d = DATASETS_DIR / ds_id
        d.mkdir(parents=True)
        ext = Path(original_filename or "").suffix.lower() or ".bin"
        (d / f"original{ext}").write_bytes(original_bytes)
        _write_parquet(df, d / "current.parquet")
        sheets = _xlsx_sheets(original_bytes) if ext == ".xlsx" else None
        meta = {
            "id": ds_id,
            "name": (name or "").strip() or Path(original_filename or "").stem or "未命名数据集",
            "original_filename": original_filename or "",
            "created_at": _now(),
            "updated_at": _now(),
            "rows": int(len(df)),
            "cols": int(df.shape[1]),
            "columns": _columns_info(df),
            "sheets": sheets or [],
            "history": [
                {
                    "time": _now(),
                    "action": action,
                    "detail": detail or f"{original_filename}（{len(df)} 行 × {df.shape[1]} 列）",
                }
            ],
        }
        _write_meta(d, meta)
    logger.info("数据集已创建 id=%s name=%s rows=%d cols=%d", ds_id, meta["name"], meta["rows"], meta["cols"])
    return ds_id


def _xlsx_sheets(raw: bytes):
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, keep_links=False)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return None


def _sniff_csv(src: Path) -> tuple:
    """嗅探大 CSV 的编码与分隔符（只读首部样本）。"""
    sample = src.read_bytes()[: 1 << 20]
    enc = "latin-1"
    for cand in ("utf-8-sig", "gbk"):
        try:
            sample.decode(cand)
            enc = cand
            break
        except UnicodeDecodeError:
            continue
    text = sample.decode(enc, errors="replace")
    lines = text.splitlines()
    head = lines[0] if lines else ""
    sep, best_n = ",", head.count(",")
    for cand in (";", "\t", "|"):
        n = head.count(cand)
        if n > best_n:
            sep, best_n = cand, n
    return enc, sep


def create_dataset_stream(name, src: Path, original_filename: str) -> str:
    """大 CSV 流式建集：分块读取直接写 Parquet，内存只驻留一个 chunk。

    分隔符误判 / 跨 chunk 类型漂移等任何异常都整体回退到全量解析路径（正确性优先）。
    """
    ds_id = _new_ds_id()
    d = DATASETS_DIR / ds_id
    d.mkdir(parents=True)
    try:
        enc, sep = _sniff_csv(src)
        import pyarrow as pa
        import pyarrow.parquet as pq

        rows, cols, columns, writer = 0, 0, [], None
        schema = None
        try:
            with open(src, "r", encoding=enc, newline="") as f:
                for chunk in pd.read_csv(f, sep=sep, chunksize=STREAM_CHUNK_ROWS):
                    if writer is None:
                        if chunk.shape[1] == 1 and any(s in str(chunk.columns[0]) for s in ",;\t|"):
                            raise ValueError(f"分隔符嗅探失败（表头未按 {sep!r} 切开）")
                        cols = int(chunk.shape[1])
                        columns = _columns_info(chunk)
                        schema = pa.Schema.from_pandas(chunk, preserve_index=False)
                        writer = pq.ParquetWriter(d / "current.parquet", schema)
                    else:
                        for c in chunk.columns:  # 跨 chunk 类型漂移：int→float 等安全提升
                            if chunk[c].dtype != schema.field(str(c)).type.to_pandas_dtype():
                                try:
                                    chunk[c] = chunk[c].astype(schema.field(str(c)).type.to_pandas_dtype())
                                except (ValueError, TypeError):
                                    raise ValueError(f"列 {c} 跨块类型冲突，回退全量解析")
                    writer.write_table(pa.Table.from_pandas(chunk, schema=schema, preserve_index=False))
                    rows += int(len(chunk))
            if writer is None:
                raise ValueError("文件解析后没有任何数据行")
        finally:
            # 句柄不关：Windows 上 rmtree 删不掉被占用的 parquet，静默泄漏孤儿目录
            if writer is not None:
                writer.close()

        # 先复制原始文件（保留 src 供失败回退），全部成功后才删 src
        ext = Path(original_filename or "").suffix.lower() or ".bin"
        shutil.copy2(src, d / f"original{ext}")
        meta = {
            "id": ds_id,
            "name": (name or "").strip() or Path(original_filename or "").stem or "未命名数据集",
            "original_filename": original_filename or "",
            "created_at": _now(),
            "updated_at": _now(),
            "rows": rows,
            "cols": cols,
            "columns": columns,
            "sheets": [],
            "history": [
                {
                    "time": _now(),
                    "action": "上传（流式）",
                    "detail": f"{original_filename}（{rows} 行 × {cols} 列，分块写入 Parquet）",
                }
            ],
        }
        with _lock:
            _write_meta(d, meta)
        src.unlink(missing_ok=True)  # 成功后才消费掉源文件（失败路径保留供回退/排查）
        logger.info("流式建集完成 id=%s rows=%d cols=%d", ds_id, rows, cols)
        return ds_id
    except Exception as e:
        logger.warning("流式建集回退全量解析（%s：%s）", type(e).__name__, e)
        shutil.rmtree(d, ignore_errors=True)
        raw = src.read_bytes()
        df = parse_upload(original_filename or "", raw)
        ds_id = create_dataset(name, df, original_filename or "", raw)
        src.unlink(missing_ok=True)  # 回退成功同样消费掉源文件（解析失败则保留）
        return ds_id


def list_datasets() -> list:
    out = []
    if DATASETS_DIR.is_dir():
        for d in DATASETS_DIR.iterdir():
            if d.is_dir() and (d / "meta.json").exists():
                try:
                    out.append(_read_meta(d))
                except (json.JSONDecodeError, OSError) as e:
                    logger.warning("跳过损坏的数据集目录 %s：%s", d.name, e)
                    continue
    out.sort(key=lambda m: (m.get("updated_at", ""), m.get("id", "")), reverse=True)
    return out


def get_meta(ds_id: str) -> dict:
    with _lock:
        return _read_meta(_ds_dir(ds_id))


def load_df(ds_id: str) -> pd.DataFrame:
    return pd.read_parquet(current_path(ds_id))


def save_df(ds_id: str, df: pd.DataFrame, action: str, detail: str = "") -> dict:
    with _lock:
        d = _ds_dir(ds_id)
        cur = d / "current.parquet"
        if not cur.exists():
            _migrate_legacy(d)
        if cur.exists():
            shutil.copy2(cur, d / "prev.parquet")  # 撤销快照：只保留最近一步
        _write_parquet(df, cur)
        meta = _read_meta(d)
        meta["rows"] = int(len(df))
        meta["cols"] = int(df.shape[1])
        meta["columns"] = _columns_info(df)
        meta["updated_at"] = _now()
        meta["history"].append({"time": _now(), "action": action, "detail": detail})
        meta["history"] = meta["history"][-MAX_HISTORY:]
        _write_meta(d, meta)
        return meta


def undo_dataset(ds_id: str) -> dict:
    """撤销最近一次修改（清洗/变换/回滚），恢复到上一步的数据。"""
    with _lock:
        d = _ds_dir(ds_id)
        prev = d / "prev.parquet"
        if prev.exists():
            df_prev = pd.read_parquet(prev)
            prev.unlink()  # 快照用完即删：只保留一级撤销
        else:
            legacy = d / "prev.pkl"  # 迁移期兼容：旧版 pickle 快照
            if not legacy.exists():
                raise DatasetNotFound(f"{ds_id}:no-undo")
            df_prev = pd.read_pickle(legacy)  # 必须转换格式，绝不能把 pickle 字节直接拷成 parquet
            legacy.unlink()
        _write_parquet(df_prev, d / "current.parquet")
        (d / "current.pkl").unlink(missing_ok=True)  # 迁移期旧版残留清理
        df = df_prev
        meta = _read_meta(d)
        undone = meta["history"].pop() if len(meta["history"]) > 1 else None
        meta["rows"] = int(len(df))
        meta["cols"] = int(df.shape[1])
        meta["columns"] = _columns_info(df)
        meta["updated_at"] = _now()
        meta["history"].append(
            {
                "time": _now(),
                "action": "撤销",
                "detail": f"撤销了「{(undone or {}).get('action', '上一步')}」",
            }
        )
        _write_meta(d, meta)
        return meta


def rename_dataset(ds_id: str, name: str) -> dict:
    with _lock:
        d = _ds_dir(ds_id)
        meta = _read_meta(d)
        meta["name"] = name.strip() or meta["name"]
        _write_meta(d, meta)
        return meta


def delete_dataset(ds_id: str) -> None:
    with _lock:
        shutil.rmtree(_ds_dir(ds_id))
    logger.info("数据集已删除 id=%s", ds_id)


def reset_dataset(ds_id: str) -> dict:
    """回滚：用原始上传文件重建当前工作副本。"""
    with _lock:
        d = _ds_dir(ds_id)
        originals = [p for p in d.glob("original.*")]
        if not originals:
            raise DatasetNotFound(ds_id)
        p = originals[0]
        df = parse_upload(p.name, p.read_bytes())
        return save_df(ds_id, df, "回滚", "恢复到上传时的原始数据")


def import_sheet(ds_id: str, sheet_name: str) -> str:
    """从原始 xlsx 的其他工作表导入为新数据集。"""
    with _lock:
        d = _ds_dir(ds_id)
        originals = [p for p in d.glob("original.xlsx")]
        if not originals:
            raise DatasetNotFound(f"{ds_id}:not-xlsx")
        p = originals[0]
        sheets = _xlsx_sheets(p.read_bytes()) or []
        if sheet_name not in sheets:
            raise DatasetNotFound(f"{ds_id}:sheet-missing:{sheet_name}")
        df = parse_upload(p.name, p.read_bytes(), sheet_name=sheet_name)
        if df.empty:
            raise ValueError(f"工作表 [{sheet_name}] 没有数据")
        meta = _read_meta(d)
        return create_dataset(
            f"{meta['name']}-{sheet_name}", df, f"{meta.get('original_filename', 'data.xlsx')}#{sheet_name}", p.read_bytes()
        )


def find_original_file(ds_id: str) -> Path:
    d = _ds_dir(ds_id)
    originals = [p for p in d.glob("original.*")]
    if not originals:
        raise DatasetNotFound(ds_id)
    return originals[0]
